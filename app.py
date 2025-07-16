import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
import streamlit as st
from io import BytesIO

# تعريف الثوابت والألوان
HEADER_FILL = PatternFill(start_color="7030A0", end_color="7030A0", fill_type="solid")
LIGHT_PURPLE_FILL = PatternFill(start_color="E4D9F5", end_color="E4D9F5", fill_type="solid")
THIN_BORDER = Border(left=Side(style='thin'), right=Side(style='thin'),
                    top=Side(style='thin'), bottom=Side(style='thin'))
CENTER_ALIGN = Alignment(horizontal='center', vertical='center')
HEADER_FONT = Font(bold=True, color="FFFFFF")

def calculate_all_metrics(df_raw_overall, df_raw_urdu, target_aht):
    """
    حساب جميع المقاييس المطلوبة من البيانات الخام
    """
    metrics = {}
    
    try:
        # 1. حساب AHT (متوسط وقت التعامل) لكل TL
        aht_scores = df_raw_overall.groupby('AD')['AH'].mean().to_dict()
        
        # 2. حساب محادثات العربية والأردية
        arabic_chats = df_raw_overall[df_raw_overall['AO'] == 'Arabic'].groupby('AD').size().to_dict()
        urdu_chats = df_raw_urdu.groupby('AD').size().to_dict()
        
        # 3. حساب Tenured AHT (للإنتاج فقط)
        tenured_aht = df_raw_overall[df_raw_overall['AF'] == 'Production'].groupby('AD')['AH'].mean().to_dict()
        
        # 4. حساب Nesting AHT
        nesting_aht = df_raw_overall[df_raw_overall['AF'] == 'Nesting'].groupby('AD')['AH'].mean().to_dict()
        
        # 5. حساب FRT (متوسط وقت الرد الأول)
        frt_scores = df_raw_overall.groupby('AD')['H'].mean().to_dict()
        
        # 6. حساب Releasing, ZTP, Missed
        releasing_counts = df_raw_overall[df_raw_overall['AI'] == 'Releasing'].groupby('AD').size().to_dict()
        ztp_counts = df_raw_overall[df_raw_overall['AJ'] == 'ZTP'].groupby('AD').size().to_dict()
        missed_counts = df_raw_overall[df_raw_overall['AK'] == 'Missed'].groupby('AD').size().to_dict()
        
        # 7. حساب Pass/Fail
        pass_counts = df_raw_overall[df_raw_overall['AC'] == 'Pass'].groupby('AD').size().to_dict()
        fail_counts = df_raw_overall[df_raw_overall['AC'] == 'Fail'].groupby('AD').size().to_dict()
        
        # 8. حساب حسب البلدان (EG, JO, BH, AE, QA, KW, OM)
        countries = ['EG', 'JO', 'BH', 'AE', 'QA', 'KW', 'OM']
        country_metrics = {}
        for country in countries:
            country_data = df_raw_overall[df_raw_overall['C'] == country]
            country_metrics[country] = country_data.groupby('AD')['AH'].mean().to_dict()
        
        # تجميع كل المقاييس في قاموس واحد
        for tl in df_raw_overall['AD'].unique():
            if pd.isna(tl):
                continue
                
            metrics[tl] = {
                'urdu': urdu_chats.get(tl, 0),
                'arabic': arabic_chats.get(tl, 0),
                'overall_aht': aht_scores.get(tl, 0),
                'tenured_aht': tenured_aht.get(tl, 0),
                'nesting_aht': nesting_aht.get(tl, 0),
                'var_from_target': aht_scores.get(tl, 0) - target_aht if tl in aht_scores else 0,
                'frt': frt_scores.get(tl, 0),
                'releasing': releasing_counts.get(tl, 0),
                'ztp': ztp_counts.get(tl, 0),
                'missed': missed_counts.get(tl, 0),
                'total_chats': arabic_chats.get(tl, 0) + urdu_chats.get(tl, 0),
                'pass': pass_counts.get(tl, 0),
                'fail': fail_counts.get(tl, 0),
                'readiness': (pass_counts.get(tl, 0) / (arabic_chats.get(tl, 0) + urdu_chats.get(tl, 0))) * 100 
                             if (arabic_chats.get(tl, 0) + urdu_chats.get(tl, 0)) > 0 else 0,
                'eg': country_metrics['EG'].get(tl, 0),
                'jo': country_metrics['JO'].get(tl, 0),
                'bh': country_metrics['BH'].get(tl, 0),
                'ae': country_metrics['AE'].get(tl, 0),
                'qa': country_metrics['QA'].get(tl, 0),
                'kw': country_metrics['KW'].get(tl, 0),
                'om': country_metrics['OM'].get(tl, 0)
            }
            
    except Exception as e:
        st.error(f"حدث خطأ في حساب المقاييس: {str(e)}")
        return {}
    
    return metrics

def create_full_report(metrics, target_aht):
    """
    إنشاء التقرير الكامل مع جميع الأقسام
    """
    try:
        wb = Workbook()
        if 'Sheet' in wb.sheetnames:
            del wb['Sheet']
        
        # 1. إنشاء ورقة View
        view_sheet = wb.create_sheet("View")
        
        # أعمدة ورقة View
        view_columns = [
            "TL", "Urdu", "Arabic", "Over all AHT Score", "Tenured AHT", "Nesting AHT",
            "# Chats Arabic", "# Chats Urdu", "Var From Target", "Status", "Readiness",
            "FRT", "EG", "JO", "BH", "AE", "QA", "KW", "OM", "Releasing", "ZTP",
            "Missed", "", "Chats", "Pass", "Fail"
        ]
        
        # إضافة العناوين
        for col_num, header in enumerate(view_columns, 1):
            cell = view_sheet.cell(row=1, column=col_num, value=header)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.border = THIN_BORDER
            cell.alignment = CENTER_ALIGN
        
        # تعبئة البيانات لكل TL
        row_num = 2
        for tl, data in metrics.items():
            # بيانات أساسية
            view_sheet.cell(row=row_num, column=1, value=tl)  # TL Name
            
            # Urdu
            view_sheet.cell(row=row_num, column=2, value=data['urdu'] if data['urdu'] > 0 else "-")
            
            # Arabic
            view_sheet.cell(row=row_num, column=3, value=data['arabic'] if data['arabic'] > 0 else "-")
            
            # Overall AHT Score
            view_sheet.cell(row=row_num, column=4, value=round(data['overall_aht'], 2) if data['overall_aht'] > 0 else "-")
            
            # Tenured AHT
            view_sheet.cell(row=row_num, column=5, value=round(data['tenured_aht'], 2) if data['tenured_aht'] > 0 else "-")
            
            # Nesting AHT
            view_sheet.cell(row=row_num, column=6, value=round(data['nesting_aht'], 2) if data['nesting_aht'] > 0 else "-")
            
            # Chats Arabic
            view_sheet.cell(row=row_num, column=7, value=data['arabic'])
            
            # Chats Urdu
            view_sheet.cell(row=row_num, column=8, value=data['urdu'])
            
            # Var From Target
            var_value = round(data['var_from_target'], 2) if isinstance(data['var_from_target'], (int, float)) else "-"
            view_sheet.cell(row=row_num, column=9, value=var_value)
            
            # Status
            status = "Achieved" if data['var_from_target'] <= 0 or data['var_from_target'] == "-" else "Not Achieved"
            view_sheet.cell(row=row_num, column=10, value=status)
            
            # Readiness
            readiness_value = f"{data['readiness']:.2f}%" if data['readiness'] > 0 else "-"
            view_sheet.cell(row=row_num, column=11, value=readiness_value)
            
            # FRT
            view_sheet.cell(row=row_num, column=12, value=round(data['frt'], 2) if data['frt'] > 0 else "-")
            
            # Countries (EG, JO, BH, AE, QA, KW, OM)
            view_sheet.cell(row=row_num, column=13, value=round(data['eg'], 2) if data['eg'] > 0 else "-")  # EG
            view_sheet.cell(row=row_num, column=14, value=round(data['jo'], 2) if data['jo'] > 0 else "-")  # JO
            view_sheet.cell(row=row_num, column=15, value=round(data['bh'], 2) if data['bh'] > 0 else "-")  # BH
            view_sheet.cell(row=row_num, column=16, value=round(data['ae'], 2) if data['ae'] > 0 else "-")  # AE
            view_sheet.cell(row=row_num, column=17, value=round(data['qa'], 2) if data['qa'] > 0 else "-")  # QA
            view_sheet.cell(row=row_num, column=18, value=round(data['kw'], 2) if data['kw'] > 0 else "-")  # KW
            view_sheet.cell(row=row_num, column=19, value=round(data['om'], 2) if data['om'] > 0 else "-")  # OM
            
            # Releasing, ZTP, Missed
            view_sheet.cell(row=row_num, column=20, value=data['releasing'])
            view_sheet.cell(row=row_num, column=21, value=data['ztp'])
            view_sheet.cell(row=row_num, column=22, value=data['missed'])
            
            # Chats, Pass, Fail
            view_sheet.cell(row=row_num, column=24, value=data['total_chats'])
            view_sheet.cell(row=row_num, column=25, value=data['pass'])
            view_sheet.cell(row=row_num, column=26, value=data['fail'])
            
            row_num += 1
        
        # حفظ الملف في الذاكرة
        output = BytesIO()
        wb.save(output)
        return output.getvalue()
        
    except Exception as e:
        st.error(f"حدث خطأ أثناء إنشاء التقرير: {str(e)}")
        return None

def main():
    st.title("نظام تحليل أداء خدمة العملاء المتكامل")
    
    # رفع الملفات
    st.subheader("رفع ملفات البيانات المطلوبة")
    raw_overall_file = st.file_uploader("ملف البيانات الكلية (Raw Overall)", type=['xlsx', 'csv'])
    raw_urdu_file = st.file_uploader("ملف بيانات الأردية (Raw Urdu)", type=['xlsx', 'csv'])
    target_aht = st.number_input("هدف AHT المطلوب (بالثواني)", value=300, min_value=0)
    
    if st.button("إنشاء التقرير"):
        if raw_overall_file and raw_urdu_file:
            try:
                # قراءة الملفات
                df_raw_overall = pd.read_excel(raw_overall_file) if raw_overall_file.name.endswith('.xlsx') else pd.read_csv(raw_overall_file)
                df_raw_urdu = pd.read_excel(raw_urdu_file) if raw_urdu_file.name.endswith('.xlsx') else pd.read_csv(raw_urdu_file)
                
                # حساب جميع المقاييس
                with st.spinner('جاري معالجة البيانات...'):
                    metrics = calculate_all_metrics(df_raw_overall, df_raw_urdu, target_aht)
                
                if metrics:
                    # إنشاء التقرير
                    with st.spinner('جاري إنشاء التقرير...'):
                        report_data = create_full_report(metrics, target_aht)
                    
                    if report_data:
                        st.success("تم إنشاء التقرير بنجاح!")
                        
                        # زر التنزيل
                        st.download_button(
                            label="⬇️ تنزيل التقرير",
                            data=report_data,
                            file_name="AHT_Full_Report.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                        )
                
            except Exception as e:
                st.error(f"حدث خطأ: {str(e)}")
        else:
            st.warning("الرجاء رفع ملفات البيانات المطلوبة")

if __name__ == "__main__":
    main()
